import streamlit as st
import pandas as pd
import re
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import numbers
from io import BytesIO

# Mapping dictionaries (unchanged)
OH_map = {'G': 'Go', '4': '4 OOS', 'R': 'Red', '5': '5 OOS', 'I': 'Rip', '2': '2', 'Y': 'Boy'}
MB_map = {'3': '3', '1': '1/Fix', 'A': 'A', 'B': 'Push A', 'C': 'C/Slide', '2': '2', 'G': 'Go', 'R': 'Red', 'E': 'None'}
OPP_map = {'G': 'Go', '4': '4 OOS', 'R': 'Red', '5': '5 OOS', 'I': 'Rip', '2': '2', 'Y': 'Boy', 'A': 'A', 'S': 'Dump'}
BR_map = {'8': 'Bic/Pipe', '7': 'Gap', '9': 'Sky/D', 'W': 'A', 'M': 'MB', '0': 'None'}
rotation_mapping = {
    '*z1': 'Rotation 1', '*z6': 'Rotation 2', '*z5': 'Rotation 3',
    '*z4': 'Rotation 4', '*z3': 'Rotation 5', '*z2': 'Rotation 6'
}

# List of teams (up to 35, starting with Stanford)
TEAMS = [
    "Stanford University", "Team 2", "Team 3", "Team 4", "Team 5",
    "Team 6", "Team 7", "Team 8", "Team 9", "Team 10",
    "Team 11", "Team 12", "Team 13", "Team 14", "Team 15",
    "Team 16", "Team 17", "Team 18", "Team 19", "Team 20",
    "Team 21", "Team 22", "Team 23", "Team 24", "Team 25",
    "Team 26", "Team 27", "Team 28", "Team 29", "Team 30",
    "Team 31", "Team 32", "Team 33", "Team 34", "Team 35"
]

# Helper Functions (unchanged except where noted)
def parse_dvw_header(content):
    lines = content.split('\n')
    match_date = "01.01"
    home_team = "Unknown Home"
    away_team = "Unknown Away"
    for i, line in enumerate(lines):
        if line.strip().startswith('[3MATCH]'):
            try:
                match_line = lines[i + 1].strip()
                if match_line:
                    date_part = match_line.split(';')[0]
                    day, month, _ = date_part.split('/')
                    match_date = f"{month}.{day}"
            except (IndexError, ValueError):
                st.warning(f"Invalid date format: {date_part}")
        elif line.strip().startswith('[3TEAMS]'):
            try:
                for j in range(i + 1, len(lines)):
                    teams_line = lines[j].strip()
                    if teams_line and not teams_line.startswith(';'):
                        teams = teams_line.split(';')
                        if len(teams) >= 3:
                            if home_team == "Unknown Home":
                                home_team = teams[1].strip()
                            else:
                                away_team = teams[1].strip()
                                break
            except IndexError:
                st.warning("Could not parse teams")
    return match_date, home_team, away_team

def extract_custom_code(line):
    pre_semicolon = line.split(';')[0]
    parts = pre_semicolon.split('~')
    return parts[-1] if parts else ""

def extract_reception(content, match_name):
    lines = content.split('\n')
    z_code = None
    receptions = []
    for line in lines:
        if z_match := re.search(r'\*z\d+', line):
            z_code = z_match.group(0)
        if z_code and (r_matches := re.findall(r'\*\d{2}R.*?;', line)):
            for r_code in r_matches:
                passer_str = r_code[1:3].strip()
                passer = int(passer_str) if passer_str.isdigit() else passer_str
                pass_grade = r_code[3] + r_code[5]
                custom_code = extract_custom_code(r_code)
                if pass_grade == 'R-' and len(custom_code) == 1 and custom_code in '45789M':
                    receptions.append((match_name, z_code, passer, pass_grade, custom_code))
                elif pass_grade in ['R#', 'R+', 'R!'] and len(custom_code) == 5 and custom_code.isalnum():
                    receptions.append((match_name, z_code, passer, pass_grade, custom_code))
    return receptions

def extract_transition(content, match_name):
    lines = content.split('\n')
    transitions = []
    z_code = None
    for i in range(2, len(lines)):
        if z_match := re.search(r'\*z\d+', lines[i-2]):
            z_code = z_match.group(0)
        if re.search(r'\*\d{2}[DF]', lines[i-2]) and (a_match := re.search(r'\*\d{2}A', lines[i])):
            attacker_str = a_match.group(0)[1:3].strip()
            attacker = int(attacker_str) if attacker_str.isdigit() else attacker_str
            custom_code = extract_custom_code(lines[i])
            if len(custom_code) == 5 and custom_code.isalnum():
                transitions.append((match_name, z_code, attacker, custom_code))
            elif len(custom_code) == 1 and custom_code in '45789M':
                transitions.append((match_name, z_code, attacker, custom_code))
    return transitions

def parse_in_system(pattern):
    if len(pattern) != 5 or not pattern.isalnum():
        return None
    oh_code, mb_code, opp_code, br_code, set_code = pattern
    oh = OH_map.get(oh_code, oh_code)
    mb = MB_map.get(mb_code, mb_code)
    opp = OPP_map.get(opp_code, opp_code)
    br = BR_map.get(br_code, br_code)
    if set_code == oh_code:
        set_to = oh
    elif set_code == mb_code:
        set_to = mb
    elif set_code == opp_code:
        set_to = opp
    elif set_code == br_code:
        set_to = br
    else:
        return None
    return {'OH': oh, 'MB': mb, 'OPP/S': opp, 'BR': br, 'set_to': set_to}

def parse_out_of_system(code):
    mapping = {'4': 'OH', '5': 'RS', 'M': 'MB', '7': 'BR', '8': 'BR', '9': 'BR'}
    return mapping.get(code)

def analyze_reception(df, rotation):
    rec_df = df[df['Reception Rotation'] == rotation]
    tallies = {}
    for cat in ['R#', 'R# or R+', 'R!']:
        if cat == 'R# or R+':
            group = rec_df[rec_df['Pass Grade'].isin(['R#', 'R+'])]
        else:
            group = rec_df[rec_df['Pass Grade'] == cat]
        cat_tallies = Counter()
        for code in group['Reception Custom Code']:
            parsed = parse_in_system(code)
            if parsed:
                key = (parsed['OH'], parsed['MB'], parsed['OPP/S'], parsed['BR'], parsed['set_to'])
                cat_tallies[key] += 1
        tallies[cat] = cat_tallies
    out_system = Counter(
        parse_out_of_system(code) for code in rec_df[rec_df['Pass Grade'] == 'R-']['Reception Custom Code']
        if parse_out_of_system(code)
    )
    tallies['R-'] = out_system
    return tallies

def analyze_transition(df, rotation):
    trans_df = df[df['Transition Rotation'] == rotation]
    in_system = Counter()
    out_system = Counter()
    for code in trans_df['Transition Custom Code']:
        if len(code) == 5:
            parsed = parse_in_system(code)
            if parsed:
                key = (parsed['OH'], parsed['MB'], parsed['OPP/S'], parsed['BR'], parsed['set_to'])
                in_system[key] += 1
        elif len(code) == 1:
            pos = parse_out_of_system(code)
            if pos:
                out_system[pos] += 1
    return in_system, out_system

def create_excel_in_memory(df, oh1_num, oh2_num, home_team):
    # (Unchanged Excel generation code - same as original)
    df['Passer #'] = df['Passer #'].astype(str)
    wb = Workbook()
    wb.remove(wb.active)

    for z_code, sheet_name in rotation_mapping.items():
        ws = wb.create_sheet(sheet_name)
        rec_df = df[df['Reception Rotation'] == z_code][[
            'Reception Match Name', 'Reception Rotation', 'Passer #', 'Pass Grade', 'Reception Custom Code'
        ]].dropna()
        ws['N1'] = 'Reception Raw Data'
        ws['N2'] = 'Match Name'
        ws['O2'] = 'Rotation'
        ws['P2'] = 'Passer #'
        ws['Q2'] = 'Pass Grade'
        ws['R2'] = 'Custom Code'
        for r, data in enumerate(rec_df.itertuples(), start=3):
            ws.cell(row=r, column=14, value=data._1)
            ws.cell(row=r, column=15, value=data._2)
            passer_val = data._3
            if isinstance(passer_val, (int, float)):
                passer_numeric = float(passer_val)
            elif isinstance(passer_val, str) and passer_val.strip().replace('-', '').replace('.', '').isdigit():
                passer_numeric = float(passer_val.strip())
            else:
                passer_numeric = None
            cell_passer = ws.cell(row=r, column=16, value=passer_numeric if passer_numeric is not None else passer_val)
            if passer_numeric is not None:
                cell_passer.number_format = "General"
            ws.cell(row=r, column=17, value=data._4)
            ws.cell(row=r, column=18, value=data._5)

        trans_df = df[df['Transition Rotation'] == z_code][[
            'Transition Match Name', 'Transition Rotation', 'Attacker #', 'Transition Custom Code'
        ]].dropna()
        ws['T1'] = 'Transition Raw Data'
        ws['T2'] = 'Match Name'
        ws['U2'] = 'Rotation'
        ws['V2'] = 'Attacker #'
        ws['W2'] = 'Custom Code'
        for r, data in enumerate(trans_df.itertuples(), start=3):
            ws.cell(row=r, column=20, value=data._1)
            ws.cell(row=r, column=21, value=data._2)
            attacker_val = data._3
            if isinstance(attacker_val, (int, float)):
                attacker_numeric = float(attacker_val)
            elif isinstance(attacker_val, str) and attacker_val.strip().replace('-', '').replace('.', '').isdigit():
                attacker_numeric = float(attacker_val.strip())
            else:
                attacker_numeric = None
            cell_attacker = ws.cell(row=r, column=22, value=attacker_numeric if attacker_numeric is not None else attacker_val)
            if attacker_numeric is not None:
                cell_attacker.number_format = "General"
            ws.cell(row=r, column=23, value=data._4)

        rec_tallies = analyze_reception(df, z_code)
        trans_in_system, trans_out_system = analyze_transition(df, z_code)

        row = 1
        for cat in ['R#', 'R# or R+', 'R!']:
            ws[f'A{row}'] = f'Reception {cat}'
            row += 1
            tallies = rec_tallies[cat]
            if tallies:
                ws[f'A{row}'] = 'OH'
                ws[f'B{row}'] = 'MB'
                ws[f'C{row}'] = 'OPP/S'
                ws[f'D{row}'] = 'BR'
                ws[f'E{row}'] = 'Set To'
                ws[f'F{row}'] = 'Count'
                row += 1
                for (oh, mb, opp, br, set_to), count in sorted(tallies.items(), key=lambda x: x[0]):
                    ws.cell(row=row, column=1, value=oh)
                    ws.cell(row=row, column=2, value=mb)
                    ws.cell(row=row, column=3, value=opp)
                    ws.cell(row=row, column=4, value=br)
                    ws.cell(row=row, column=5, value=set_to)
                    cell_count = ws.cell(row=row, column=6, value=float(count))
                    cell_count.number_format = "General"
                    row += 1
                row += 1

        ws.cell(row=row, column=1, value='Reception R-')
        row += 1
        out_system = rec_tallies['R-']
        if out_system:
            ws.cell(row=row, column=1, value='Position')
            ws.cell(row=row, column=2, value='Count')
            row += 1
            for pos, count in sorted(out_system.items(), key=lambda x: x[0]):
                ws.cell(row=row, column=1, value=pos)
                cell_count = ws.cell(row=row, column=2, value=float(count))
                cell_count.number_format = "General"
                row += 1
            row += 1

        ws.cell(row=row, column=1, value='Transition In-System')
        row += 1
        if trans_in_system:
            ws.cell(row=row, column=1, value='OH')
            ws.cell(row=row, column=2, value='MB')
            ws.cell(row=row, column=3, value='OPP/S')
            ws.cell(row=row, column=4, value='BR')
            ws.cell(row=row, column=5, value='Set To')
            ws.cell(row=row, column=6, value='Count')
            row += 1
            for (oh, mb, opp, br, set_to), count in sorted(trans_in_system.items(), key=lambda x: x[0]):
                ws.cell(row=row, column=1, value=oh)
                ws.cell(row=row, column=2, value=mb)
                ws.cell(row=row, column=3, value=opp)
                ws.cell(row=row, column=4, value=br)
                ws.cell(row=row, column=5, value=set_to)
                cell_count = ws.cell(row=row, column=6, value=float(count))
                cell_count.number_format = "General"
                row += 1
            row += 1

        ws.cell(row=row, column=1, value='Transition OOS TR')
        row += 1
        if trans_out_system:
            ws.cell(row=row, column=1, value='Position')
            ws.cell(row=row, column=2, value='Count')
            row += 1
            for pos, count in sorted(trans_out_system.items(), key=lambda x: x[0]):
                ws.cell(row=row, column=1, value=pos)
                cell_count = ws.cell(row=row, column=2, value=float(count))
                cell_count.number_format = "General"
                row += 1

        formatted_row = 1
        pattern_based = ['Reception R#', 'Reception R# or R+', 'Reception R!', 'Transition In-System']
        position_based = ['Reception R-', 'Transition OOS TR']
        category_key_map = {
            'Reception R#': 'R#', 'Reception R# or R+': 'R# or R+', 'Reception R!': 'R!',
            'Reception R-': 'R-', 'Transition In-System': 'Transition In-System', 'Transition OOS TR': 'Transition OOS TR'
        }

        pattern_data = []
        for cat in pattern_based:
            key = category_key_map[cat]
            tallies = trans_in_system if key == 'Transition In-System' else rec_tallies[key]
            for (oh, mb, opp, br, set_to), count in tallies.items():
                pattern_data.append({
                    'Category': cat, 'OH': oh, 'MB': mb, 'OPP/S': opp, 'BR': br, 'Set To': set_to, 'Count': count
                })
        pattern_df = pd.DataFrame(pattern_data)

        for cat in pattern_based:
            ws[f'H{formatted_row}'] = cat
            cat_df = pattern_df[pattern_df['Category'] == cat]
            grouped = cat_df.groupby(['OH', 'MB', 'OPP/S', 'BR'])
            for (oh, mb, opp, br), group in grouped:
                states = [oh, mb, opp, br]
                counter = {state: 0 for state in states}
                for _, row_data in group.iterrows():
                    set_to = row_data['Set To']
                    if set_to in counter:
                        counter[set_to] += row_data['Count']
                total_count = sum(counter.values())
                percentages = {state: (counter[state] / total_count) if total_count > 0 else 0.0 for state in states}
                ws.cell(row=formatted_row+1, column=9, value=states[0])
                ws.cell(row=formatted_row+1, column=10, value=states[1])
                ws.cell(row=formatted_row+1, column=11, value=states[2])
                ws.cell(row=formatted_row+1, column=12, value=states[3])
                cell_total = ws.cell(row=formatted_row+2, column=8, value=float(total_count))
                cell_total.number_format = "General"
                for col, state in enumerate(states, start=9):
                    cell_value = ws.cell(row=formatted_row+2, column=col, value=float(counter[state]))
                    cell_value.number_format = "General"
                for col, state in enumerate(states, start=9):
                    ws.cell(row=formatted_row+3, column=col, value=f"{percentages[state]:.2f}")
                formatted_row += 5

        for cat in position_based:
            ws[f'H{formatted_row}'] = cat
            key = category_key_map[cat]
            tallies = trans_out_system if key == 'Transition OOS TR' else rec_tallies[key]
            positions = ['OH', 'MB', 'OPP/S', 'BR']
            counter = {pos: tallies.get(pos, 0) for pos in positions}
            total_count = sum(counter.values())
            percentages = {pos: (counter[pos] / total_count) if total_count > 0 else 0.0 for pos in positions}
            for col, pos in enumerate(positions, start=9):
                ws.cell(row=formatted_row+1, column=col, value=pos)
            cell_total = ws.cell(row=formatted_row+2, column=8, value=float(total_count))
            cell_total.number_format = "General"
            for col, pos in enumerate(positions, start=9):
                cell_value = ws.cell(row=formatted_row+2, column=col, value=float(counter[pos]))
                cell_value.number_format = "General"
            for col, pos in enumerate(positions, start=9):
                ws.cell(row=formatted_row+3, column=col, value=f"{percentages[pos]:.2f}")
            formatted_row += 5

    ws_set_odds = wb.create_sheet("Set Odds")
    def calculate_set_odds(player_num, relevant_z_codes, rotation_labels, start_row):
        ws_set_odds[f'A{start_row}'] = f"#{player_num} - Odds of Getting Set After a #{player_num} Reception"
        ws_set_odds[f'B{start_row+1}'] = f"After #{player_num} Passed In System"
        ws_set_odds[f'F{start_row+1}'] = f"After Someone Other #{player_num} Passed In System"
        ws_set_odds[f'B{start_row+2}'] = f"#{player_num} Was Set"
        ws_set_odds[f'C{start_row+2}'] = f"#{player_num} Not Set"
        ws_set_odds[f'D{start_row+2}'] = f"#{player_num} Was Set %"
        ws_set_odds[f'F{start_row+2}'] = f"#{player_num} Was Set"
        ws_set_odds[f'G{start_row+2}'] = f"#{player_num} Not Set"
        ws_set_odds[f'H{start_row+2}'] = f"#{player_num} Was Set %"

        row = start_row + 3
        total_was_set_passer = 0
        total_not_set_passer = 0
        total_was_set_other = 0
        total_not_set_other = 0

        for rot_name, z_code in zip(rotation_labels, relevant_z_codes):
            ws_set_odds[f'A{row}'] = rot_name
            filtered_df = df[(df['Reception Rotation'] == z_code) &
                             (df['Pass Grade'].isin(['R#', 'R+', 'R!'])) &
                             (df['Reception Custom Code'].notna())].copy()
            filtered_df['parsed'] = filtered_df['Reception Custom Code'].apply(parse_in_system)
            filtered_df = filtered_df[filtered_df['parsed'].notna()]

            passer_df = filtered_df[filtered_df['Passer #'] == player_num]
            was_set = (passer_df['parsed'].apply(lambda x: x['set_to'] == x['OH'])).sum()
            not_set = len(passer_df) - was_set
            total = was_set + not_set
            percentage = was_set / total if total > 0 else 0
            ws_set_odds[f'B{row}'] = was_set
            ws_set_odds[f'C{row}'] = not_set
            cell_d = ws_set_odds[f'D{row}']
            cell_d.value = percentage
            cell_d.number_format = '0.00'

            other_df = filtered_df[filtered_df['Passer #'] != player_num]
            was_set_other = (other_df['parsed'].apply(lambda x: x['set_to'] == x['OH'])).sum()
            not_set_other = len(other_df) - was_set_other
            total_other = was_set_other + not_set_other
            percentage_other = was_set_other / total_other if total_other > 0 else 0
            ws_set_odds[f'F{row}'] = was_set_other
            ws_set_odds[f'G{row}'] = not_set_other
            cell_h = ws_set_odds[f'H{row}']
            cell_h.value = percentage_other
            cell_h.number_format = '0.00'

            total_was_set_passer += was_set
            total_not_set_passer += not_set
            total_was_set_other += was_set_other
            total_not_set_other += not_set_other
            row += 1

        ws_set_odds[f'A{row}'] = "Tot"
        total_passer = total_was_set_passer + total_not_set_passer
        percentage_passer = total_was_set_passer / total_passer if total_passer > 0 else 0
        ws_set_odds[f'B{row}'] = total_was_set_passer
        ws_set_odds[f'C{row}'] = total_not_set_passer
        cell_d_total = ws_set_odds[f'D{row}']
        cell_d_total.value = percentage_passer
        cell_d_total.number_format = '0.00'

        total_other = total_was_set_other + total_not_set_other
        percentage_other = total_was_set_other / total_other if total_other > 0 else 0
        ws_set_odds[f'F{row}'] = total_was_set_other
        ws_set_odds[f'G{row}'] = total_not_set_other
        cell_h_total = ws_set_odds[f'H{row}']
        cell_h_total.value = percentage_other
        cell_h_total.number_format = '0.00'

        return row + 2

    oh1_rotations = [('*z1', 'Rot 1'), ('*z3', 'Rot 5'), ('*z2', 'Rot 6')]
    oh1_z_codes, oh1_labels = zip(*oh1_rotations)
    next_row = calculate_set_odds(oh1_num, oh1_z_codes, oh1_labels, 1)

    oh2_rotations = [('*z6', 'Rot 2'), ('*z5', 'Rot 3'), ('*z4', 'Rot 4')]
    oh2_z_codes, oh2_labels = zip(*oh2_rotations)
    calculate_set_odds(oh2_num, oh2_z_codes, oh2_labels, next_row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

### Streamlit App
st.title("Volleyball Match Analysis")

# Get team from URL parameter if present
query_params = st.query_params
selected_team = query_params.get("team", TEAMS[0])  # Default to Stanford if no param

# Team Selection
team = st.selectbox("Select Team to Analyze", TEAMS, index=TEAMS.index(selected_team))
st.write(f"Analyzing files where **{team}** is the home team.")

# File Upload
uploaded_files = st.file_uploader(f"Upload .dvw files for {team}", type="dvw", accept_multiple_files=True)

if uploaded_files:
    # Filter files by selected home team
    file_contents = {}
    valid_files = []
    for uploaded_file in uploaded_files:
        content = uploaded_file.read().decode('ISO-8859-1')
        _, home_team, _ = parse_dvw_header(content)
        if home_team == team:
            file_contents[uploaded_file.name] = content
            valid_files.append(uploaded_file.name)
        else:
            st.warning(f"File '{uploaded_file.name}' skipped: Home team '{home_team}' does not match selected team '{team}'.")

    if not valid_files:
        st.error(f"No uploaded files have '{team}' as the home team.")
    else:
        st.success(f"Found {len(valid_files)} valid files for {team}.")

        # Extract opponents (away teams) from valid files
        opponents = set()
        for content in file_contents.values():
            _, _, away_team = parse_dvw_header(content)
            opponents.add(away_team)
        opponents = sorted(list(opponents))

        # Opponent Selection
        selected_opponents = st.multiselect("Select Opponents to Analyze", opponents, default=opponents)
        if not selected_opponents:
            st.warning("Please select at least one opponent to proceed.")
        else:
            # Process only selected opponents
            filtered_file_contents = {
                fname: content for fname, content in file_contents.items()
                if parse_dvw_header(content)[2] in selected_opponents
            }

            # Extract data
            all_receptions = []
            all_transitions = []
            for fname, content in filtered_file_contents.items():
                match_date, _, away_team = parse_dvw_header(content)
                match_name = f"{match_date} {away_team}"
                receptions = extract_reception(content, match_name)
                transitions = extract_transition(content, match_name)
                all_receptions.extend(receptions)
                all_transitions.extend(transitions)

            # Create DataFrame
            all_data = []
            max_length = max(len(all_receptions), len(all_transitions))
            for i in range(max_length):
                rec_row = list(all_receptions[i]) if i < len(all_receptions) else ["", "", "", "", ""]
                trans_row = list(all_transitions[i]) if i < len(all_transitions) else ["", "", "", ""]
                row = rec_row + ["", ""] + trans_row
                all_data.append(row)
            df = pd.DataFrame(all_data, columns=[
                "Reception Match Name", "Reception Rotation", "Passer #", "Pass Grade", "Reception Custom Code", "", "",
                "Transition Match Name", "Transition Rotation", "Attacker #", "Transition Custom Code"
            ])

            # User inputs for OH1 and OH2
            oh1_num = st.text_input("Enter the number of OH1:", "")
            oh2_num = st.text_input("Enter the number of OH2:", "")

            if oh1_num and oh2_num:
                if st.button("Generate and Download Excel File"):
                    with st.spinner("Generating Excel file..."):
                        excel_file = create_excel_in_memory(df, oh1_num, oh2_num, team)
                        st.download_button(
                            label="Download Analysis Excel",
                            data=excel_file,
                            file_name=f"{team} Analysis.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        st.success("Excel file generated successfully!")
else:
    st.info(f"Please upload .dvw files for {team} to begin analysis.")